"""
Agency / Special Motor Matrix Grid Processor
Flask backend — fully dynamic, config-driven
"""

from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
import openpyxl
import os, uuid, traceback, json
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'outputs')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

ALLOWED = {'xlsx', 'xlsb', 'xls', 'csv'}
def ok_file(f): return '.' in f and f.rsplit('.',1)[1].lower() in ALLOWED


# ══════════════════════════════════════════════════════════════
# SHEET INTROSPECTION
# ══════════════════════════════════════════════════════════════

def get_sheets(path):
    ext = path.rsplit('.',1)[1].lower()
    if ext == 'xlsb':
        import pyxlsb
        with pyxlsb.open_workbook(path) as wb:
            return wb.sheets
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    s = wb.sheetnames; wb.close(); return s


def read_sheet_headers(path, sheet, header_rows, start_row, start_col, n_preview=5):
    """
    Read multi-row headers + first n_preview data rows.
    Returns col_defs with both a human-readable display and a match_key for default lookup.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    max_col = ws.max_column
    max_row = ws.max_row

    sorted_hrs = sorted(header_rows)
    last_hr = sorted_hrs[-1]          # data-label row (e.g. row 4 = 'GWP')
    parent_hr = sorted_hrs[0]         # group row    (e.g. row 2 = 'PVT CAR(1+1)')
    sub_hr = sorted_hrs[1] if len(sorted_hrs) > 1 else sorted_hrs[0]  # row 3 = sub-cat

    # Read all header rows
    header_data = {}
    for hr in header_rows:
        header_data[hr] = [ws.cell(hr, c).value for c in range(1, max_col+1)]

    # Meta cols (1..start_col-1)
    meta_defs = []
    for c in range(1, start_col):
        last_val = header_data[last_hr][c-1]
        display  = str(last_val).strip() if last_val else f'Col{c}'
        meta_defs.append({'col_idx': c, 'display': display,
                          'labels': {str(hr): str(header_data[hr][c-1] or '') for hr in header_rows}})

    # Rate cols (start_col..max_col)
    col_defs = []
    for c in range(start_col, max_col+1):
        parent_val = str(header_data[parent_hr][c-1] or '').strip()
        sub_val    = str(header_data[sub_hr][c-1] or '').strip()
        last_val   = str(header_data[last_hr][c-1] or '').strip()

        # Human display: "PVT CAR(1+1) | DIESEL & NCB"
        if sub_val and sub_val != parent_val:
            display = f"{parent_val} | {sub_val}" if parent_val else sub_val
        elif parent_val:
            display = parent_val
        else:
            display = last_val or f'Col{c}'

        # match_key for default config lookup: (parent_biz_upper, sub_label_upper)
        match_key = (parent_val.upper(), sub_val.upper())

        col_defs.append({
            'col_idx': c,
            'display': display,
            'parent_biz': parent_val,
            'sub_label': sub_val,
            'match_key': match_key,
            'labels': {str(hr): str(header_data[hr][c-1] or '') for hr in header_rows}
        })

    # Preview rows
    all_defs = meta_defs + col_defs
    preview = []
    for r in range(start_row, min(start_row + n_preview, max_row+1)):
        rd = {}
        for cd in all_defs:
            v = ws.cell(r, cd['col_idx']).value
            rd[cd['display']] = '' if v is None else str(v)
        preview.append(rd)

    # Unique Volume Consideration values
    vc_col = next((cd['col_idx'] for cd in meta_defs if 'volume' in cd['display'].lower()), None)
    vol_values = []
    if vc_col:
        seen = set()
        for r in range(start_row, max_row+1):
            v = ws.cell(r, vc_col).value
            sv = str(v).strip() if v else ''
            if sv and sv not in seen:
                seen.add(sv); vol_values.append(sv)

    # Unique IMD Type values
    imd_col = next((cd['col_idx'] for cd in meta_defs if 'imd type' in cd['display'].lower()), None)
    imd_values = []
    if imd_col:
        seen = set()
        for r in range(start_row, max_row+1):
            v = ws.cell(r, imd_col).value
            sv = str(v).strip() if v else ''
            if sv and sv not in seen:
                seen.add(sv); imd_values.append(sv)

    wb.close()
    return {
        'meta_defs': meta_defs,
        'col_defs': col_defs,
        'preview': preview,
        'vol_values': vol_values,
        'imd_values': imd_values,
        'max_row': max_row,
        'max_col': max_col
    }


# ══════════════════════════════════════════════════════════════
# RTO LOOKUP
# ══════════════════════════════════════════════════════════════

def build_rto_index(path, sheet, header_row,
                    rto_col='RTO CODE', cluster_col='UW CLUSTER (25-26)', cat_col='PRODUCT CATEGORY'):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    headers = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(header_row, c).value
        if v: headers[str(v).strip().upper()] = c
    rtoc = headers.get(rto_col.upper())
    catc = headers.get(cat_col.upper())
    # Auto-detect cluster column: try exact match first, then any "UW CLUSTER" variant
    cluc = headers.get(cluster_col.upper())
    if not cluc:
        for hdr, idx in headers.items():
            if hdr.startswith('UW CLUSTER'):
                cluc = idx
                break
    index = {}
    all_codes = []
    for r in range(header_row+1, ws.max_row+1):
        rto = ws.cell(r, rtoc).value if rtoc else None
        clu = ws.cell(r, cluc).value if cluc else None
        cat = ws.cell(r, catc).value if catc else None
        if rto:
            all_codes.append(str(rto).strip())
            if clu and cat:
                key = (str(clu).strip().upper(), str(cat).strip().upper())
                index.setdefault(key, set()).add(str(rto).strip())
    wb.close()
    return index, sorted(set(all_codes))


def lookup_rto(index, all_codes, cluster, category, norm_map=None, fallback_to_all=True):
    def _try(clu, cat):
        return index.get((str(clu).strip().upper(), str(cat).strip().upper()))
    codes = _try(cluster, category)
    if not codes and norm_map:
        norm_cat = norm_map.get(category.strip().upper())
        if norm_cat:
            codes = _try(cluster, norm_cat)
    if not codes:
        if fallback_to_all and all_codes:
            return ','.join(all_codes)
        return 'ANY'
    return ','.join(sorted(codes))


# ══════════════════════════════════════════════════════════════
# CORE PROCESSING ENGINE
# ══════════════════════════════════════════════════════════════

def process_matrix(config, rto_index=None, all_rto_codes=None):
    path        = config['filepath']
    sheet       = config['sheet_name']
    header_rows = config.get('header_rows', [2, 3, 4])
    data_start  = config.get('data_start_row', 5)
    meta_map    = config.get('meta_col_map', {})
    col_defs    = config.get('col_defs', [])
    mode        = config.get('mode', 'special')
    ignore_vals = set(config.get('ignore_values', ['Block','NA','IRDA','MISP','SYSTEM COMMISSION']))
    irda_vals   = set(config.get('irda_values',   ['Block','NA','IRDA','MISP','SYSTEM COMMISSION']))
    irda_prct   = config.get('irda_prct_value',   '-0.1')
    irda_outgo  = config.get('irda_outgo_value',  'IRDA')
    norm_outgo  = config.get('normal_outgo_value', 'GWP')
    agent_map   = config.get('agent_group_map', {})
    vol_gwp_map = config.get('vol_gwp_map', {})
    # column_defaults: dict of {output_col_name: default_value} applied to every output row
    # when the field would otherwise be absent or empty. Set via UI "Column Defaults" section.
    col_defaults = config.get('column_defaults', {})
    std_ll      = config.get('std_gwp_ll_col',    'Total Gwp Ll*')
    std_ul      = config.get('std_gwp_ul_col',    'Total Gwp Ul*')
    prime_ll    = config.get('prime_gwp_ll_col',  'Total Gwp Ll*')
    prime_ul    = config.get('prime_gwp_ul_col',  'Total Gwp Ul*')
    agency_ll   = config.get('agency_gwp_ll_col', 'Totalgwp Keybrok Agency Ll*')
    agency_ul   = config.get('agency_gwp_ul_col', 'Totalgwp Keybrok Agency Ul*')
    static_flds = config.get('output_static_fields', {})
    rto_norm    = {k.upper():v for k,v in config.get('rto_norm_map', {
        'PCV 3W':'PCV','PCV-BUS':'PCV','PCV-TAXI':'PCV',
        'MISD':'MISC-D','MISD GARBAGE':'MISC-D','TRACTOR':'MISC-D'
    }).items()}
    skip_vol_biz= set(v.strip() for v in config.get('skip_if_vol_biz', []))
    span_out    = config.get('span_outgo_col',   'Span Outgo*')
    span_pct    = config.get('span_prct_col',    'Span Prct*')
    rto_code_c  = config.get('rto_code_col',     'Rto Code*')
    rto_clu_c   = config.get('rto_cluster_col',  'Rto Cluster*')
    parent_c    = config.get('parent_agent_col', 'Parent Agent Code*')
    primary_c   = config.get('primary_agent_col','Primary Agent Code*')
    ag_grp_c    = config.get('agent_group_col',  'Agent Group Code*')
    biz_mix_c   = config.get('biz_mix_col',      'Biz Mix*')

    # col_idx=0 means this column is absent in this sheet layout
    def ci(key):
        idx = meta_map.get(key, {}).get('col_idx', 0)
        return idx if idx and idx > 0 else None

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    max_row = ws.max_row

    def cell(r, c):
        if not c: return ''
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ''

    output_rows = []
    skipped_rows = 0
    skipped_cells = 0

    for row_idx in range(data_start, max_row + 1):
        imd_code = cell(row_idx, ci('imd_code'))
        imd_name = cell(row_idx, ci('imd_name'))
        rel_code = cell(row_idx, ci('rel_code'))
        imd_type = cell(row_idx, ci('imd_type'))
        vol_ll   = cell(row_idx, ci('vol_ll'))
        vol_ul   = cell(row_idx, ci('vol_ul'))
        vol_rem  = cell(row_idx, ci('vol_remark'))
        uw_clust = cell(row_idx, ci('uw_cluster'))

        # ── Blank row guard ──────────────────────────────────────────────────
        # col_idx=0 means absent. Only check columns that actually exist.
        # Std Grid sheets (no IMD code/name, just cluster+vol) set imd cols to 0,
        # so we fall back to checking uw_cluster / vol_ll as identity.
        # A row is skipped only when ALL configured identity columns are empty.
        identity_map = {
            'imd_code': imd_code, 'imd_name': imd_name,
            'uw_cluster': uw_clust, 'vol_ll': vol_ll,
        }
        configured_identity = [k for k in ['imd_code','imd_name','uw_cluster','vol_ll']
                                if ci(k) is not None]
        if configured_identity:
            if not any(identity_map[k] for k in configured_identity):
                skipped_rows += 1; continue
        # If zero identity cols configured at all: skip only via cell loop below

        # Skip header-bleed rows: vol_ll contains a label string like "Agency"
        if skip_vol_biz and vol_ll in skip_vol_biz:
            skipped_rows += 1; continue

        for cd in col_defs:
            col_idx   = cd.get('col_idx', 0)
            biz_mix   = cd.get('biz_mix_output', '')
            extra     = cd.get('extra_fields', {})
            rto_cat   = cd.get('rto_category', '')

            if not col_idx:
                continue

            raw_val = cell(row_idx, col_idx)

            # Always skip truly empty cells — no value means no commission defined
            if not raw_val:
                skipped_cells += 1; continue

            # ── Normal mode ──────────────────────────────────────────────────
            # If the cell value is in the ignore list (Block, NA, IRDA, MISP,
            # SYSTEM COMMISSION etc.) → skip THIS cell only, move to next column.
            # The row continues; other columns for this same IMD row are still processed.
            if mode == 'normal':
                if raw_val in ignore_vals:
                    skipped_cells += 1; continue

            # ── Special mode ─────────────────────────────────────────────────
            # If the cell value is in the irda trigger list → do NOT skip.
            # Instead produce an output row with Span Outgo = IRDA, Span Prct = irda_prct.
            # Only skip if value is empty (handled above).

            # Build output row
            out = {}

            # Static fields first
            out.update(static_flds)

            # Agent Group Code
            ag_code = agent_map.get(imd_type, '')
            if ag_code:
                out[ag_grp_c] = ag_code

            # GWP LL/UL
            gwp_ll_c, gwp_ul_c = _resolve_gwp_cols(
                vol_rem, imd_type, vol_gwp_map,
                std_ll, std_ul, prime_ll, prime_ul, agency_ll, agency_ul
            )
            if gwp_ll_c: out[gwp_ll_c] = vol_ll
            if gwp_ul_c: out[gwp_ul_c] = vol_ul

            # Agent codes
            out[parent_c] = imd_code
            if rel_code:
                out[primary_c] = rel_code

            # RTO Cluster
            out[rto_clu_c] = uw_clust

            # Span Outgo / Prct — core logic
            if mode == 'special':
                if raw_val in irda_vals:
                    out[span_out] = irda_outgo
                    out[span_pct] = irda_prct
                else:
                    out[span_out] = norm_outgo
                    out[span_pct] = raw_val
            else:
                out[span_out] = norm_outgo
                out[span_pct] = raw_val

            # Biz Mix
            out[biz_mix_c] = biz_mix

            # Extra fields (type of business, fuel type, CC, GWP, vehicle age etc.)
            out.update(extra)

            # RTO codes
            if rto_index and uw_clust and rto_cat:
                rto = lookup_rto(rto_index, all_rto_codes or [], uw_clust, rto_cat, rto_norm)
            else:
                rto = 'ANY'
            out[rto_code_c] = rto

            # Apply column_defaults: fill any key not already set (or set to '')
            for def_col, def_val in col_defaults.items():
                if def_col not in out or out[def_col] == '':
                    out[def_col] = def_val

            output_rows.append(out)

    wb.close()
    return output_rows, skipped_rows, skipped_cells


def _resolve_gwp_cols(vol_rem, imd_type, vol_gwp_map,
                      std_ll, std_ul, prime_ll, prime_ul, agency_ll, agency_ul):
    vr = (vol_rem or '').strip().lower()
    for k, v in vol_gwp_map.items():
        if k.lower() == vr:
            return v.get('ll_col', ''), v.get('ul_col', '')
    if vr == 'std-grid':
        if imd_type and 'prime' in imd_type.lower():
            return prime_ll, prime_ul
        return agency_ll, agency_ul
    return std_ll, std_ul


# ══════════════════════════════════════════════════════════════
# DEFAULT COL CONFIG (from JS scripts, keyed by parent_biz+sub_label)
# ══════════════════════════════════════════════════════════════

# Key: (PARENT_BIZ_UPPER, SUB_LABEL_UPPER)  — matches the Excel header structure
DEFAULT_COL_CONFIG_KEYED = {
    # GCV
    ('GCV <=2.5 T',  'GCV <=2.5 T'):     {'biz_mix_output':'GCV <=2.5 T',   'rto_category':'GCV', 'extra_fields':{'Gross Vehicle Weight Ll*':'0','Gross Vehicle Weight Ul*':'2500.001'}},
    ('GCV 2.5T - 2.8T','GCV 2.5T - 2.8T'):{'biz_mix_output':'GCV 2.5T - 2.8T','rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'2500.001','Gross Vehicle Weight Ul*':'2800.001'}},
    ('GCV 2.8T - 3.5T','GCV 2.8T - 3.5T'):{'biz_mix_output':'GCV 2.8T - 3.5T','rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'2800.001','Gross Vehicle Weight Ul*':'3500.001'}},
    ('GCV 3.5T - 7.5T','GCV 3.5T - 7.5T'):{'biz_mix_output':'GCV 3.5T - 7.5T','rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'3500.001','Gross Vehicle Weight Ul*':'7500.001'}},
    ('GCV 7.5T - 12T','GCV 7.5T - 12T'):  {'biz_mix_output':'GCV 7.5T - 12T', 'rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'7500.001','Gross Vehicle Weight Ul*':'12000.001'}},
    ('GCV 12T - 20T','GCV 12T-20T AGE<5'):{'biz_mix_output':'GCV 12T - 20T',  'rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'12000.001','Gross Vehicle Weight Ul*':'20000.001','Vehicle Age Ul*':'5'}},
    ('GCV 12T - 20T','GCV 12T-20T AGE>=5'):{'biz_mix_output':'GCV 12T - 20T', 'rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'12000.001','Gross Vehicle Weight Ul*':'20000.001','Vehicle Age Ll*':'5'}},
    ('GCV 20T - 40T','GCV 20T-40T AGE<5'):{'biz_mix_output':'GCV 20T - 40T',  'rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'20000.001','Gross Vehicle Weight Ul*':'40000.001','Vehicle Age Ul*':'5'}},
    ('GCV 20T - 40T','GCV 20T-40T AGE>=5'):{'biz_mix_output':'GCV 20T - 40T', 'rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'20000.001','Gross Vehicle Weight Ul*':'40000.001','Vehicle Age Ll*':'5'}},
    ('GCV > 40T',    'GCV > 40T'):         {'biz_mix_output':'GCV > 40T',      'rto_category':'GCV','extra_fields':{'Gross Vehicle Weight Ll*':'40000.001'}},
    ('GCV-3W',       'GCV-3W'):            {'biz_mix_output':'GCV-3W',         'rto_category':'GCV','extra_fields':{}},
    # PCV-3W
    ('PCV-3W','PCV 3W ELECTRIC'): {'biz_mix_output':'PCV-3W','rto_category':'PCV 3W','extra_fields':{'Fuel Type*':'Electric'}},
    ('PCV-3W','PCV 3W NEW'):      {'biz_mix_output':'PCV-3W','rto_category':'PCV 3W','extra_fields':{'Type Of Business*':'New','Fuel Type*':'Petrol, Diesel'}},
    ('PCV-3W','PCV 3W OLD'):      {'biz_mix_output':'PCV-3W','rto_category':'PCV 3W','extra_fields':{'Type Of Business*':'Renewal, Roll Over','Fuel Type*':'Petrol, Diesel'}},
    # PCV-BUS
    ('PCV-BUS','PCV-BUS_OTHER'):  {'biz_mix_output':'PCV-BUS','rto_category':'PCV-BUS','extra_fields':{'Bus Type*':'Other Bus'}},
    ('PCV-BUS','PCV-BUS_SCHOOL'): {'biz_mix_output':'PCV-BUS','rto_category':'PCV-BUS','extra_fields':{'Bus Type*':'School Bus'}},
    # PCV-TAXI
    ('PCV-TAXI','PCV-TAXI'):      {'biz_mix_output':'PCV-TAXI','rto_category':'PCV-TAXI','extra_fields':{}},
    # TRACTOR
    ('TRACTOR','TRACTOR NEW'):    {'biz_mix_output':'TRACTOR','rto_category':'Tractor','extra_fields':{'Type Of Business*':'New'}},
    ('TRACTOR','TRACTOR OLD'):    {'biz_mix_output':'TRACTOR','rto_category':'Tractor','extra_fields':{'Type Of Business*':'Renewal, Roll Over'}},
    # CE
    ('CE','CE-CONSTRUCTION EQ'):  {'biz_mix_output':'CE-CONSTRUCTION EQ','rto_category':'MISD','extra_fields':{}},
    ('CE','MISD GARBAGE'):        {'biz_mix_output':'MISD GARBAGE',      'rto_category':'MISD Garbage','extra_fields':{}},
    ('CE','MISC-D GARBAGE'):      {'biz_mix_output':'MISD GARBAGE',      'rto_category':'MISD Garbage','extra_fields':{}},
    ('CE','MISC-D OTHERS'):       {'biz_mix_output':'CE-CONSTRUCTION EQ','rto_category':'MISD',         'extra_fields':{}},
    ('CE','HARVESTER NEW'):       {'biz_mix_output':'HARVESTER NEW',     'rto_category':'MISD','extra_fields':{'Type Of Business*':'New'}},
    ('CE','HARVESTER OLD'):       {'biz_mix_output':'HARVESTER OLD',     'rto_category':'MISD','extra_fields':{'Type Of Business*':'Renewal, Roll Over'}},
    # 2W
    ('2W','<75CC'):     {'biz_mix_output':'2W','rto_category':'2W','extra_fields':{'Cubic Capacity Ul*':'76'}},
    ('2W','75-150CC'):  {'biz_mix_output':'2W','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'76','Cubic Capacity Ul*':'151'}},
    ('2W','150-350CC'): {'biz_mix_output':'2W','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'151','Cubic Capacity Ul*':'351'}},
    ('2W','>350CC'):    {'biz_mix_output':'2W','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'351'}},
    ('2W','SCOOTER'):   {'biz_mix_output':'2W','rto_category':'2W','extra_fields':{'Two Wheeler Category*':'Scooter'}},
    # 2W(1+1)
    ('2W(1+1)','<75CC'):    {'biz_mix_output':'2W(1+1)','rto_category':'2W','extra_fields':{'Cubic Capacity Ul*':'76'}},
    ('2W(1+1)','75-150CC'): {'biz_mix_output':'2W(1+1)','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'76','Cubic Capacity Ul*':'151'}},
    ('2W(1+1)','150-350CC'):{'biz_mix_output':'2W(1+1)','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'151','Cubic Capacity Ul*':'351'}},
    ('2W(1+1)','>350CC'):   {'biz_mix_output':'2W(1+1)','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'351'}},
    ('2W(1+1)','SCOOTER'):  {'biz_mix_output':'2W(1+1)','rto_category':'2W','extra_fields':{'Two Wheeler Category*':'Scooter'}},
    # 2W(1+5)
    ('2W(1+5)','<75CC'):    {'biz_mix_output':'2W(1+5)','rto_category':'2W','extra_fields':{'Cubic Capacity Ul*':'76'}},
    ('2W(1+5)','75-150CC'): {'biz_mix_output':'2W(1+5)','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'76','Cubic Capacity Ul*':'151'}},
    ('2W(1+5)','150-350CC'):{'biz_mix_output':'2W(1+5)','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'151','Cubic Capacity Ul*':'351'}},
    ('2W(1+5)','>350CC'):   {'biz_mix_output':'2W(1+5)','rto_category':'2W','extra_fields':{'Cubic Capacity Ll*':'351'}},
    ('2W(1+5)','SCOOTER'):  {'biz_mix_output':'2W(1+5)','rto_category':'2W','extra_fields':{'Two Wheeler Category*':'Scooter'}},
    # PVT CAR(1+1)
    ('PVT CAR(1+1)','DIESEL & NCB'):      {'biz_mix_output':'PVT CAR(1+1)','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Diesel','Ncb Ll*':'1'}},
    ('PVT CAR(1+1)','DIESEL & ZERO NCB'): {'biz_mix_output':'PVT CAR(1+1)','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Diesel','Ncb Ul*':'1'}},
    ('PVT CAR(1+1)','PETROL & NCB'):      {'biz_mix_output':'PVT CAR(1+1)','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Petrol','Ncb Ll*':'1'}},
    ('PVT CAR(1+1)','PETROL & ZERO NCB'): {'biz_mix_output':'PVT CAR(1+1)','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Petrol','Ncb Ul*':'1'}},
    # PVT CAR(1+3)
    ('PVT CAR(1+3)','DIESEL'): {'biz_mix_output':'PVT CAR(1+3)','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Diesel'}},
    ('PVT CAR(1+3)','PETROL'): {'biz_mix_output':'PVT CAR(1+3)','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Petrol'}},
    # PVT CAR(3+3)
    ('PVT CAR(3+3)','ROLLOVER DIESEL & NCB'):      {'biz_mix_output':'PVT CAR(3+3)','rto_category':'Pvt Car','extra_fields':{'Type Of Business*':'Roll Over','Fuel Type*':'Diesel','Ncb Ll*':'1'}},
    ('PVT CAR(3+3)','ROLLOVER DIESEL & ZERO NCB'): {'biz_mix_output':'PVT CAR(3+3)','rto_category':'Pvt Car','extra_fields':{'Type Of Business*':'Roll Over','Fuel Type*':'Diesel','Ncb Ul*':'1'}},
    ('PVT CAR(3+3)','ROLLOVER PETROL & NCB'):      {'biz_mix_output':'PVT CAR(3+3)','rto_category':'Pvt Car','extra_fields':{'Type Of Business*':'Roll Over','Fuel Type*':'Petrol','Ncb Ll*':'1'}},
    ('PVT CAR(3+3)','ROLLOVER PETROL & ZERO NCB'): {'biz_mix_output':'PVT CAR(3+3)','rto_category':'Pvt Car','extra_fields':{'Type Of Business*':'Roll Over','Fuel Type*':'Petrol','Ncb Ul*':'1'}},
    ('PVT CAR(3+3)','NEW DIESEL'): {'biz_mix_output':'PVT CAR(3+3)','rto_category':'Pvt Car','extra_fields':{'Type Of Business*':'New','Fuel Type*':'Diesel'}},
    ('PVT CAR(3+3)','NEW PETROL'): {'biz_mix_output':'PVT CAR(3+3)','rto_category':'Pvt Car','extra_fields':{'Type Of Business*':'New','Fuel Type*':'Petrol'}},
    # PVT CAR (standalone)
    ('PVT CAR','PETROL & NCB'):      {'biz_mix_output':'PVT CAR','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Petrol','Ncb Ll*':'1'}},
    ('PVT CAR','PETROL & ZERO NCB'): {'biz_mix_output':'PVT CAR','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Petrol','Ncb Ul*':'1'}},
    ('PVT CAR','DIESEL & NCB'):      {'biz_mix_output':'PVT CAR','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Diesel','Ncb Ll*':'1'}},
    ('PVT CAR','DIESEL & ZERO NCB'): {'biz_mix_output':'PVT CAR','rto_category':'Pvt Car','extra_fields':{'Fuel Type*':'Diesel','Ncb Ul*':'1'}},
}

def get_default_for_col(parent_biz, sub_label):
    """Match a col to the default config using both parent and sub label (case-insensitive)."""
    key = (parent_biz.strip().upper(), sub_label.strip().upper())
    d = DEFAULT_COL_CONFIG_KEYED.get(key)
    if d: return d
    # Fallback: match by sub_label only (when parent==sub, many GCV cols)
    for (p, s), v in DEFAULT_COL_CONFIG_KEYED.items():
        if s == key[1]: return v
    return None



# ══════════════════════════════════════════════════════════════
# COLUMN TRANSFORMS & OUTPUT FORMAT
# ══════════════════════════════════════════════════════════════

TRANSFORM_OPS = {
    'multiply':  lambda v, n: float(v) * n,
    'divide':    lambda v, n: float(v) / n if n != 0 else float(v),
    'add':       lambda v, n: float(v) + n,
    'subtract':  lambda v, n: float(v) - n,
    'round':     lambda v, n: round(float(v), int(n)),
}

def apply_col_transforms(df, transforms):
    """
    transforms: list of {col, op, value}
    Applies in order. Non-numeric cells are left as-is.
    Returns modified df (copy).
    """
    if not transforms:
        return df
    df = df.copy()
    for t in transforms:
        col = t.get('col', '').strip()
        op  = t.get('op', '').strip().lower()
        try:
            num = float(t.get('value', 0))
        except (TypeError, ValueError):
            continue
        if not col or op not in TRANSFORM_OPS or col not in df.columns:
            continue
        fn = TRANSFORM_OPS[op]
        def safe_apply(v):
            try:
                result = fn(v, num)
                # Return int if result is a whole number
                if isinstance(result, float) and result == int(result):
                    return int(result)
                return result
            except (TypeError, ValueError):
                return v
        df[col] = df[col].apply(safe_apply)
    return df


def apply_output_format(df, output_format):
    """
    output_format: list of {col, rename} in desired order.
    Only columns listed are included; rename is optional.
    If output_format is empty, return df unchanged.
    """
    if not output_format:
        return df
    cols_in_df = set(df.columns)
    selected = [f for f in output_format if f.get('col','').strip() in cols_in_df]
    if not selected:
        return df
    col_order = [f['col'].strip() for f in selected]
    df = df[col_order].copy()
    rename_map = {f['col'].strip(): f['rename'].strip()
                  for f in selected if f.get('rename','').strip()}
    if rename_map:
        df = df.rename(columns=rename_map)
    return df

# ══════════════════════════════════════════════════════════════
# API ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/')
def index(): return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload():
    try:
        if 'file' not in request.files: return jsonify({'error':'No file'}),400
        f = request.files['file']
        if not f.filename or not ok_file(f.filename): return jsonify({'error':'Invalid file type'}),400
        sid = str(uuid.uuid4())[:8]
        fn  = secure_filename(f.filename)
        sp  = os.path.join(UPLOAD_DIR, f"{sid}_{fn}"); f.save(sp)
        rto_path = None
        if 'rto_file' in request.files:
            rf = request.files['rto_file']
            if rf.filename and ok_file(rf.filename):
                rn = secure_filename(rf.filename)
                rto_path = os.path.join(UPLOAD_DIR, f"{sid}_rto_{rn}"); rf.save(rto_path)
        sheets = get_sheets(sp)
        return jsonify({'session_id':sid,'filepath':sp,'rto_filepath':rto_path,'sheets':sheets,'filename':fn})
    except Exception as e:
        return jsonify({'error':str(e)}),500


@app.route('/api/inspect', methods=['POST'])
def inspect():
    try:
        d = request.json
        result = read_sheet_headers(
            d['filepath'], d['sheet_name'],
            d.get('header_rows',[2,3,4]),
            d.get('data_start_row',5),
            d.get('start_col',13)
        )
        # Attach default config to each col_def so frontend can show pre-filled values
        enriched_cols = []
        for cd in result['col_defs']:
            def_cfg = get_default_for_col(cd['parent_biz'], cd['sub_label'])
            enriched_cols.append({
                'col_idx':    cd['col_idx'],
                'display':    cd['display'],
                'parent_biz': cd['parent_biz'],
                'sub_label':  cd['sub_label'],
                'biz_mix_output': def_cfg['biz_mix_output'] if def_cfg else cd['display'],
                'rto_category':   def_cfg['rto_category']   if def_cfg else '',
                'extra_fields':   def_cfg['extra_fields']   if def_cfg else {},
                'matched_default': bool(def_cfg)
            })
        return jsonify({
            'col_defs':   enriched_cols,
            'meta_defs':  result['meta_defs'],
            'preview':    result['preview'],
            'vol_values': result['vol_values'],
            'imd_values': result['imd_values'],
            'max_row':    result['max_row'],
            'max_col':    result['max_col']
        })
    except Exception as e:
        return jsonify({'error':str(e),'trace':traceback.format_exc()}),500


@app.route('/api/process', methods=['POST'])
def process():
    try:
        d = request.json
        config = d.get('config', {})
        config['filepath'] = d['filepath']

        rto_index = None; all_rto = []
        # Use dedicated RTO file if uploaded; otherwise fall back to the main grid file
        # (RTO sheet is often inside the same workbook as the grid)
        rto_src = d.get('rto_filepath')
        if not rto_src or not os.path.exists(str(rto_src)):
            rto_src = d.get('filepath')   # same workbook fallback
        if rto_src and os.path.exists(str(rto_src)):
            try:
                rto_index, all_rto = build_rto_index(
                    rto_src,
                    d.get('rto_sheet', 'RTO Vs Cluster (New)'),
                    d.get('rto_header_row', 2),
                    d.get('rto_col', 'RTO CODE'),
                    d.get('rto_cluster_col', 'UW CLUSTER (26-27)'),
                    d.get('rto_cat_col', 'PRODUCT CATEGORY')
                )
            except Exception as rto_err:
                print(f'[WARN] RTO index build failed: {rto_err}')

        rows, skipped_rows, skipped_cells = process_matrix(config, rto_index, all_rto)

        if not rows:
            return jsonify({'error':f'No output rows generated. {skipped_rows} source rows skipped (blank/header). Check sheet config.'}),400

        df = pd.DataFrame(rows)

        # Apply column transforms (e.g. multiply Span Prct* by 100)
        col_transforms = d.get('col_transforms', [])
        if col_transforms:
            df = apply_col_transforms(df, col_transforms)

        # Apply output format (column selection + ordering + optional rename)
        output_format = d.get('output_format', [])
        if output_format:
            df = apply_output_format(df, output_format)

        sid  = d.get('session_id','x')
        fn   = f"{sid}_{secure_filename(d.get('output_name','agency_output'))}.csv"
        op   = os.path.join(OUTPUT_DIR, fn)
        df.to_csv(op, index=False)

        return jsonify({
            'success':True,'output_path':op,'output_filename':fn,
            'rows':len(df),'cols':len(df.columns),
            'skipped_rows':skipped_rows,
            'skipped_cells':skipped_cells,
            'preview':df.head(5).fillna('').astype(str).to_dict('records'),
            'columns':list(df.columns)
        })
    except Exception as e:
        return jsonify({'error':str(e),'trace':traceback.format_exc()}),500


@app.route('/api/download/<filename>')
def download(filename):
    p = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(p): return jsonify({'error':'Not found'}),404
    return send_file(p, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    app.run(debug=True, port=5051, host='0.0.0.0')
